import openpyxl
from openpyxl.utils import get_column_letter

from django.core.management import BaseCommand

from main.models import User
from main.models.site import BuyObject, ObjectField


class Command(BaseCommand):
    help = 'load xlsx avito database'

    def handle(self, *args, **options):
        workbook = openpyxl.load_workbook("main/management/commands/avito.xlsx")
        # Выбираем активный лист
        sheet = workbook.active
        # Собираем все столбцы имеющиеся в объявлении
        columns = {}
        for i in range(1, sheet.max_column):
            if sheet[f'{get_column_letter(i)}1'].value not in [
                'Id', 'ImageUrls', 'ManagerName', 'EMail', 'DevelopmentsSpec', 'AvitoStatus', 'AvitoDateEnd',
                'CompanyName', 'ListingFee']:
                columns[sheet[f'{get_column_letter(i)}1'].value] = get_column_letter(i)
        print(columns)
        for c in columns:
            print(c)
        # Iterate the loop to read the cell values
        for i in range(2, sheet.max_row + 1):
            # Модель BuyObject
            try:
                BuyObject.objects.get(avito_id=sheet[f'{columns["AvitoId"]}{i}'].value)
                continue
            except BuyObject.DoesNotExist:
                buy_obj = BuyObject()
                buy_obj.avito_id = sheet[f'{columns["AvitoId"]}{i}'].value
            try:
                buy_obj.user = User.objects.get(phone=sheet[f'{columns["ContactPhone"]}{i}'].value)
            except User.DoesNotExist:
                buy_obj.user = None
            buy_obj.save()
            for c in columns:
                if c in ['AvitoId', 'ContactPhone']:
                    continue
                if sheet[f'{columns[c]}{i}'].value:
                    try:
                        obj_field = ObjectField.objects.get(obj=buy_obj, name=c)
                    except ObjectField.DoesNotExist:
                        obj_field = ObjectField()
                        obj_field.obj = buy_obj
                        obj_field.name = c
                    obj_field.value = sheet[f'{columns[c]}{i}'].value
                    obj_field.save()
