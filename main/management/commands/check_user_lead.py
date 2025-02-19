from asgiref.sync import async_to_sync
from django.db.models import Q
from django.urls import reverse
from openpyxl.workbook import Workbook

from django.core.management import BaseCommand

from main.models import Comment
from main.tasks import send_to_group


class Command(BaseCommand):
    help = 'fix sources dates for statistic (use only ONCE)'

    def handle(self, *args, **options):
        for n in ['Архипова Юлия', 'Трещалин Александр']:
            comments = Comment.objects.filter(Q(text__icontains=n) & Q(text__icontains='=> <b>Комбаров Антон</b>'), type='lead')
            for comment in comments:
                print(n, comment.type, comment.item_id)
        exit()
        leads = Lead.objects.filter(processed__in=['redo', 'not'], leadsource__source__name='База для обзвона')
        wb = Workbook()
        ws = wb.active
        ws.title = 'База клиентов'
        ws['A1'] = 'ФИО'
        ws['B1'] = 'Ссылка'
        ws['C1'] = 'Телефон'
        ws['D1'] = 'Ответственный'
        ws['E1'] = 'Сделка'
        ws['F1'] = 'Ссылка'
        ws['G1'] = 'Способ оплаты'
        ws['H1'] = 'Дата брони'
        ws['I1'] = 'Стадия продажи'
        for lead in leads:
            for deal in lead.deal_set.all():
                ws.append(
                    [lead.__str__(),
                     f'https://crm.квартиры36.рф{reverse("main:lead-page", kwargs={"lead_id": lead.pk})}',
                     lead.phone, lead.responsible.__str__(), deal.__str__(),
                     f'https://crm.квартиры36.рф{reverse("main:deal-page", kwargs={"deal_id": deal.pk})}',
                     deal.paytype_text(), deal.reserved, deal.stage.__str__() if deal.stage else ''])
        wb.save(filename=f'/var/www/crm/База клиентов_64.xlsx')
        with open(f'/var/www/crm/База клиентов_64.xlsx', 'rb') as file:
            async_to_sync(send_to_group)(1123662127, '*Отфильтрованная база клиентов*', file)
