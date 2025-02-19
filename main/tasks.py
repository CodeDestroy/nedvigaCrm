from datetime import datetime

from asgiref.sync import async_to_sync
from django.conf import settings
from celery import shared_task
from django.db.models import Q
from django.urls import reverse
from openpyxl.workbook import Workbook
from telegram import Bot
from telegram.constants import ParseMode

from main.models import Lead, Comment


@shared_task()
def send_lead_report(request):
    leads = Lead.objects.all()
    if user := request.get('user'):
        leads = leads.filter(responsible__pk=user)
    if text := request.get('text'):
        leads = leads.filter(
            Q(surname__icontains=text) | Q(name__icontains=text) | Q(phone__icontains=text) |
            Q(pk__in=Comment.objects.filter(text__icontains=text, type='lead').values_list('item_id', flat=True)))
    if request.get('reserve'):
        leads = leads.filter(deal__reserved__isnull=False)
    if request.get('spam'):
        leads = leads.exclude(warm__in=['not_client', 'not_call'])
    if request.get('no_source'):
        leads = leads.filter(leadsource__isnull=True)
    if source := request.get('source'):
        leads = leads.filter(leadsource__source__pk=source)
    if warm := request.get('warm'):
        leads = leads.filter(warm=warm)
    if processed := request.get('processed'):
        leads = leads.filter(processed=processed)
    if paytype := request.get('paytype'):
        leads = leads.filter(Q(deal__paytype=paytype) | Q(questions__payment=paytype))
    if frm := request.get('from'):
        leads = leads.filter(deal__frm=frm)
    if decoration := request.get('decoration'):
        leads = leads.filter(questions__decoration=decoration)
    if marital := request.get('marital'):
        leads = leads.filter(questions__marital_status=marital)
    if purpose := request.get('purpose'):
        leads = leads.filter(questions__purpose=purpose)
    if birth := request.get('birth'):
        leads = leads.filter(questions__birth_date=datetime.strptime(birth, '%Y-%m-%d').date())
    if reserved := request.get('reserved'):
        leads = leads.filter(deal__reserved=datetime.strptime(reserved, '%Y-%m-%d').date())
    wb = Workbook()
    ws = wb.active
    ws.title = 'База клиентов'
    ws['A1'] = 'ФИО'
    ws['B1'] = 'Ссылка'
    ws['C1'] = 'Телефон'
    ws['D1'] = 'Ответственный'
    ws['E1'] = 'Сделка'
    ws['F1'] = 'Ссылка'
    ws['G1'] = 'Способ оплаты'
    ws['H1'] = 'Дата брони'
    ws['I1'] = 'Дата сделки'
    ws['J1'] = 'Стадия продажи'
    for lead in leads:
        for deal in lead.deal_set.all():
            ws.append(
                [lead.__str__(), f'https://crm.квартиры36.рф{reverse("main:lead-page", kwargs={"lead_id": lead.pk})}',
                 lead.phone, lead.responsible.__str__(), deal.__str__(),
                 f'https://crm.квартиры36.рф{reverse("main:deal-page", kwargs={"deal_id": deal.pk})}',
                 deal.paytype_text(), deal.reserved, deal.sell_date, deal.stage.__str__() if deal.stage else ''])
    wb.save(filename=f'/var/www/crm/База клиентов_{request["user_id"]}.xlsx')
    with open(f'/var/www/crm/База клиентов_{request["user_id"]}.xlsx', 'rb') as file:
        async_to_sync(send_to_group)(request['telegram_id'], '*Отфильтрованная база клиентов*', file)


async def send_to_group(telegram_id, text, file):
    async with Bot(settings.TELEGRAM_BOT_KEY) as bot:
        # admin user id = 286896743
        await bot.send_document(telegram_id, caption=text, parse_mode=ParseMode.MARKDOWN_V2, document=file)
